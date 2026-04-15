#!/usr/bin/env python3
"""
tools/import_xlsx.py
Import EMF IG Tracking spreadsheet (Sheet1) into the flat scrape_data table.
Fully idempotent — safe to re-run on updated files.

Usage:
    python tools/import_xlsx.py <path_to_xlsx>

Example:
    python tools/import_xlsx.py ".tmp/all_scrape_data.xlsx"

Required packages:
    pip install openpyxl supabase python-dotenv

Required .env:
    SUPABASE_URL=https://xxxx.supabase.co
    SUPABASE_SERVICE_KEY=eyJ...   (service_role key, NOT anon key)
"""

import sys
import os
import time
import math
from datetime import datetime, date

# ── dependency check ──────────────────────────────────────────────────────────
try:
    import openpyxl
    from dotenv import load_dotenv
    from supabase import create_client, Client
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Run: pip install openpyxl supabase python-dotenv")
    sys.exit(1)

# ── config ────────────────────────────────────────────────────────────────────
load_dotenv()
SUPABASE_URL = os.getenv('SUPABASE_URL', '')
SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_KEY', '')

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Error: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set in .env")
    sys.exit(1)

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
BATCH_SIZE = 100


# ── helpers ───────────────────────────────────────────────────────────────────
def parse_date(val) -> str | None:
    if not val:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()[:10]

    # ISO format first
    try:
        return datetime.strptime(s, '%Y-%m-%d').strftime('%Y-%m-%d')
    except ValueError:
        pass

    # Slash-separated: DD/MM/YYYY (UK format used in all_scrape_data.xlsx)
    parts = s.split('/')
    if len(parts) == 3:
        try:
            a, b, yr = int(parts[0]), int(parts[1]), parts[2]
            if a > 12:
                # a cannot be a month → must be DD/MM/YYYY
                return date(int(yr), b, a).strftime('%Y-%m-%d')
            elif b > 12:
                # b cannot be a month → must be MM/DD/YYYY (day in second position)
                return date(int(yr), a, b).strftime('%Y-%m-%d')
            else:
                # Ambiguous (both ≤ 12): UK format → DD/MM first, fall back to MM/DD
                for month, day in [(b, a), (a, b)]:
                    try:
                        return date(int(yr), month, day).strftime('%Y-%m-%d')
                    except ValueError:
                        continue
        except (ValueError, TypeError):
            pass

    return None


def to_float(val) -> float | None:
    if val is None:
        return None
    try:
        f = float(val)
        return f if math.isfinite(f) else None
    except (ValueError, TypeError):
        return None


def to_int(val) -> int | None:
    f = to_float(val)
    return int(f) if f is not None else None


MAX_RETRIES = 5

def upsert_with_retry(batch: list):
    """Upsert one batch via RPC (function-level 120s timeout) with exponential backoff."""
    for attempt in range(MAX_RETRIES):
        try:
            supabase.rpc('upsert_scrape_batch', {'batch': batch}).execute()
            return
        except Exception as e:
            msg = str(e)
            # Retryable: gateway errors, rate limits, transient network issues
            if any(code in msg for code in ('502', '503', '429', 'timeout', 'ConnectionError')):
                wait = 2 ** attempt  # 1s, 2s, 4s, 8s, 16s
                print(f"\n   [retry {attempt+1}/{MAX_RETRIES}] transient error, waiting {wait}s … ({msg[:80]})")
                time.sleep(wait)
            else:
                raise  # non-retryable — propagate immediately
    raise RuntimeError(f"Batch upsert failed after {MAX_RETRIES} attempts")


def flush(rows: list) -> int:
    if not rows:
        return 0
    # Deduplicate by row_key — keep the last occurrence so the most recent
    # data wins when the same (post_id, scrape_date) appears more than once.
    deduped = {r['row_key']: r for r in rows}
    unique = list(deduped.values())
    for i in range(0, len(unique), BATCH_SIZE):
        batch = unique[i:i + BATCH_SIZE]
        upsert_with_retry(batch)
    return len(unique)


# ── main ──────────────────────────────────────────────────────────────────────
def main(xlsx_path: str):
    print(f"\n── Reading {xlsx_path}")
    # read_only=True streams row-by-row — fast and low RAM even for 200k+ rows
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    if 'Sheet1' not in wb.sheetnames:
        print("Error: 'Sheet1' not found in workbook.")
        sys.exit(1)

    ws = wb['Sheet1']

    buffer   = []
    total    = 0
    upserted = 0
    skipped  = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        if total % 1_000 == 0:
            print(f"   {total:,} rows scanned, {upserted:,} upserted …", flush=True)

        # ── unpack columns (order matches sheet) ─────────────────────────────
        # Pad short rows so column access never raises IndexError
        row = tuple(row) + (None,) * max(0, 30 - len(row))
        post_url      = row[0]
        raw_post_id   = row[1]
        post_ts       = parse_date(row[2])
        content_type  = row[3]
        caption       = row[4]
        verticals     = row[5]
        comments      = to_float(row[6])
        likes         = to_float(row[7])
        views         = to_float(row[8]) or 0.0
        scrape_date   = parse_date(row[9])
        username      = str(row[10]).strip() if row[10] else None
        # row[11] = model_code       — dropped from DB, skip
        model_stage   = row[12]
        x_score       = to_float(row[13])
        followers     = to_float(row[14])
        employee      = row[15] if row[15] and str(row[15]).strip() != '-' else None
        tracking_link = row[16]
        acc_no        = to_int(row[17])
        trial         = row[18]
        # row[19] = views_gained     — dropped, computed via LAG() in DB
        # row[20] = datetrackinglink — dropped, skip
        # row[21] = thumbnail        — dropped, skip
        shared_feed   = bool(row[22]) if row[22] is not None else None
        # row[23] = curr_vert        — dropped, derived from latest verticals per username
        user_id       = to_int(row[24])
        talent_mgr    = row[25]
        # row[26] = curr_user        — dropped, derived from latest username per user_id
        # row[27] = curr_model       — dropped, derived from account_model table
        # row[28] = curr_tm          — dropped, derived from latest talent_manager per user_id
        team          = row[29]

        # ── require at minimum a username and scrape date ─────────────────────
        if not username or not scrape_date:
            skipped += 1
            continue

        # ── build stable upsert key ───────────────────────────────────────────
        post_id = str(raw_post_id).strip() if raw_post_id else None
        if post_id and post_id.lower() not in ('none', ''):
            row_key = f"{post_id}__{scrape_date}"
        else:
            post_id = None
            row_key = f"acct__{username}__{scrape_date}"

        buffer.append({
            'row_key':           row_key,
            'post_url':          str(post_url) if post_url else None,
            'post_id':           post_id,
            'post_timestamp':    post_ts,
            'content_type':      str(content_type) if content_type else None,
            'caption':           str(caption) if caption else None,
            'verticals':         str(verticals) if verticals else None,
            'comments':          comments,
            'likes':             likes,
            'views':             views,
            'scrape_date':       scrape_date,
            'username':          username,
            'model_stage_name':  str(model_stage) if model_stage else None,
            'x_score':           x_score,
            'followers':         followers,
            'employee':          str(employee) if employee else None,
            'tracking_link':     str(tracking_link) if tracking_link else None,
            'acc_no':            acc_no,
            'trial':             str(trial) if trial else None,
            'is_shared_to_feed': shared_feed,
            'user_id':           user_id,
            'talent_manager':    str(talent_mgr) if talent_mgr else None,
            'team':              str(team) if team else None,
        })

        # ── flush every BATCH_SIZE rows to keep memory flat ───────────────────
        if len(buffer) >= BATCH_SIZE:
            upserted += flush(buffer)
            buffer.clear()
            time.sleep(0.2)  # let free-tier IOPS recover between batches

    # ── flush remainder ───────────────────────────────────────────────────────
    upserted += flush(buffer)

    print(f"\n── Import complete.")
    print(f"   {total:,} rows read | {skipped:,} skipped | {upserted:,} upserted")

    # Refresh both materialized views so the dashboard picks up new data.
    # Order matters: mv_post_latest first, then mv_account_median (which reads from it).
    # The RPC sets statement_timeout = 600s internally to bypass the default 3s cap.
    print("── Refreshing materialized views …", end=' ', flush=True)
    try:
        supabase.rpc('refresh_materialized_views').execute()
        print("done\n")
    except Exception as e:
        print(f"WARNING: refresh failed ({e}).")
        print("   Run manually in Supabase SQL editor:")
        print("   SELECT refresh_materialized_views();\n")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1])
